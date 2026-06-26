from __future__ import annotations

import csv

from rdkit import Chem
from rdkit.Chem import AllChem

from ezscreen.results.export import export_sdf, export_xlsx

_HEADERS = ["ligand", "score", "name", "smiles", "LE", "BEI"]
_ROWS = [
    {"ligand": "lig1", "score": "-9.5", "name": "alpha", "smiles": "CCO", "LE": "0.42", "BEI": "20.1"},
    {"ligand": "lig2", "score": "-7.2", "name": "beta",  "smiles": "c1ccccc1", "LE": "0.61", "BEI": "15.0"},
]


def _write_scores(path):
    with path.open("w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=_HEADERS)
        w.writeheader()
        w.writerows(_ROWS)


def _write_poses(path):
    writer = Chem.SDWriter(str(path))
    for row in _ROWS:
        mol = Chem.AddHs(Chem.MolFromSmiles(row["smiles"]))
        AllChem.EmbedMolecule(mol, randomSeed=1)
        mol.SetProp("lig_id", row["ligand"])
        writer.write(mol)
    writer.close()


def test_export_xlsx_round_trip(tmp_path):
    from openpyxl import load_workbook

    scores = tmp_path / "scores.csv"
    _write_scores(scores)
    out = export_xlsx(scores, tmp_path / "hits.xlsx")
    assert out.exists()

    wb = load_workbook(out)
    ws = wb.active
    header = [c.value for c in ws[1]]
    assert header == _HEADERS
    assert ws.max_row == len(_ROWS) + 1

    # score column is written as a number, not a string
    score_idx = _HEADERS.index("score") + 1
    assert ws.cell(row=2, column=score_idx).value == -9.5


def test_export_xlsx_limit(tmp_path):
    from openpyxl import load_workbook

    scores = tmp_path / "scores.csv"
    _write_scores(scores)
    out = export_xlsx(scores, tmp_path / "hits.xlsx", limit=1)

    wb = load_workbook(out)
    ws = wb.active
    assert ws.max_row == 2  # header + 1 data row
    # the kept row is the best (first) one
    name_idx = _HEADERS.index("name") + 1
    assert ws.cell(row=2, column=name_idx).value == "alpha"


def test_export_sdf_limit(tmp_path):
    scores = tmp_path / "scores.csv"
    poses = tmp_path / "poses.sdf"
    _write_scores(scores)
    _write_poses(poses)

    out = export_sdf(poses, scores, tmp_path / "hits.sdf", limit=1)
    mols = [m for m in Chem.SDMolSupplier(str(out)) if m is not None]
    assert len(mols) == 1
    assert mols[0].GetProp("name") == "alpha"


_VARIANT_HEADERS = ["ligand", "score", "name", "smiles", "LE", "BEI"]
_VARIANT_ROWS = [
    {"ligand": "lig1", "score": "-9.5", "name": "alpha_v2", "smiles": "CCO", "LE": "0.42", "BEI": "20.1"},
    {"ligand": "lig2", "score": "-8.0", "name": "alpha_v1", "smiles": "CCO", "LE": "0.40", "BEI": "19.0"},
    {"ligand": "lig3", "score": "-7.2", "name": "beta",     "smiles": "c1ccccc1", "LE": "0.61", "BEI": "15.0"},
]


def _write_variant_scores(path):
    with path.open("w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=_VARIANT_HEADERS)
        w.writeheader()
        w.writerows(_VARIANT_ROWS)


def test_export_xlsx_collapse(tmp_path):
    from openpyxl import load_workbook

    scores = tmp_path / "scores.csv"
    _write_variant_scores(scores)
    out = export_xlsx(scores, tmp_path / "hits.xlsx", collapse=True)

    wb = load_workbook(out)
    ws = wb.active
    header = [c.value for c in ws[1]]
    assert "forms" in header
    # alpha (2 forms) collapses to its best, beta stays → 2 data rows
    assert ws.max_row == 3
    name_idx = header.index("name") + 1
    forms_idx = header.index("forms") + 1
    assert ws.cell(row=2, column=name_idx).value == "alpha_v2"  # best form kept
    assert ws.cell(row=2, column=forms_idx).value == 2


def test_export_sdf_collapse(tmp_path):
    scores = tmp_path / "scores.csv"
    poses = tmp_path / "poses.sdf"
    _write_variant_scores(scores)

    writer = Chem.SDWriter(str(poses))
    for row in _VARIANT_ROWS:
        mol = Chem.AddHs(Chem.MolFromSmiles(row["smiles"]))
        AllChem.EmbedMolecule(mol, randomSeed=1)
        mol.SetProp("lig_id", row["ligand"])
        writer.write(mol)
    writer.close()

    out = export_sdf(poses, scores, tmp_path / "hits.sdf", collapse=True)
    mols = [m for m in Chem.SDMolSupplier(str(out)) if m is not None]
    # only the best form of alpha (lig1) + beta (lig3)
    names = {m.GetProp("name") for m in mols}
    assert names == {"alpha_v2", "beta"}


def test_export_sdf_attaches_scores(tmp_path):
    scores = tmp_path / "scores.csv"
    poses = tmp_path / "poses.sdf"
    _write_scores(scores)
    _write_poses(poses)

    out = export_sdf(poses, scores, tmp_path / "hits.sdf")
    assert out.exists()

    mols = [m for m in Chem.SDMolSupplier(str(out)) if m is not None]
    assert len(mols) == len(_ROWS)

    by_name = {m.GetProp("name"): m for m in mols}
    assert set(by_name) == {"alpha", "beta"}
    assert by_name["alpha"].GetProp("score") == "-9.5"
    assert by_name["alpha"].GetProp("LE") == "0.42"
    assert by_name["alpha"].HasProp("smiles")
