from __future__ import annotations

import csv
from pathlib import Path

_AMBER = "FFE3B341"


def _detect_score_col(headers: list[str]) -> str | None:
    for h in headers:
        if "score" in h.lower() or "affinity" in h.lower():
            return h
    return headers[-1] if headers else None


def export_xlsx(scores_csv: Path, out_xlsx: Path, limit: int | None = None, collapse: bool = False) -> Path:
    """Write a styled Excel hit list from a merged scores.csv.

    limit caps the number of rows written (top N, since scores.csv is sorted best-first);
    None writes every row. collapse keeps only the best-scoring enumerated form per source
    molecule and adds a "forms" column with how many forms collapsed.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    with scores_csv.open(newline="") as f:
        reader = csv.DictReader(f)
        headers = list(reader.fieldnames or [])
        rows = list(reader)

    if collapse:
        from ezscreen.results.variants import collapse_variants
        rows = collapse_variants(rows)
        for r in rows:
            r["forms"] = r.pop("variant_count", 1)
        headers = headers + ["forms"]

    if limit is not None and limit > 0:
        rows = rows[:limit]

    wb = Workbook()
    ws = wb.active
    ws.title = "Hits"

    bold = Font(bold=True)
    amber = PatternFill(start_color=_AMBER, end_color=_AMBER, fill_type="solid")

    ws.append(headers)
    for cell in ws[1]:
        cell.font = bold

    score_col = _detect_score_col(headers)
    score_idx = headers.index(score_col) + 1 if score_col in headers else None
    le_idx = headers.index("LE") + 1 if "LE" in headers else None

    for row in rows:
        ws.append([row.get(h, "") for h in headers])
        r = ws.max_row
        if score_idx is not None:
            cell = ws.cell(row=r, column=score_idx)
            try:
                cell.value = float(cell.value)
                cell.number_format = "0.000"
            except (TypeError, ValueError):
                pass
        if le_idx is not None:
            cell = ws.cell(row=r, column=le_idx)
            try:
                if float(cell.value) > 0.5:
                    cell.fill = amber
            except (TypeError, ValueError):
                pass

    ws.freeze_panes = "A2"
    for i, h in enumerate(headers, 1):
        width = max(len(str(h)), *(len(str(row.get(h, ""))) for row in rows)) if rows else len(str(h))
        ws.column_dimensions[get_column_letter(i)].width = min(max(width + 2, 8), 60)

    wb.save(out_xlsx)
    return out_xlsx


def export_sdf(poses_sdf: Path, scores_csv: Path, out_sdf: Path, limit: int | None = None, collapse: bool = False) -> Path:
    """Write poses.sdf with score/LE/BEI and other scores.csv fields attached as SD properties.

    limit keeps only poses among the top N ligands of scores.csv (best-first); None keeps
    every pose present locally. collapse keeps only the best enumerated form per source
    molecule. Poses beyond the run's returned cap simply aren't available.
    """
    from rdkit import Chem

    index: dict[str, dict] = {}
    id_col = "ligand"
    with scores_csv.open(newline="") as f:
        reader = csv.DictReader(f)
        headers = list(reader.fieldnames or [])
        if headers:
            id_col = headers[0]
        all_rows = list(reader)
    for row in all_rows:
        index[row.get(id_col, "")] = row

    ranked_rows = all_rows
    if collapse:
        from ezscreen.results.variants import collapse_variants
        ranked_rows = collapse_variants(all_rows)
    ordered_ids = [r.get(id_col, "") for r in ranked_rows]

    allowed: set[str] | None = None
    if collapse:
        allowed = set(ordered_ids)
    if limit is not None and limit > 0:
        allowed = set(ordered_ids[:limit])

    supplier = Chem.SDMolSupplier(str(poses_sdf), removeHs=False, sanitize=True)
    writer = Chem.SDWriter(str(out_sdf))
    try:
        for i, mol in enumerate(supplier):
            if mol is None:
                continue
            lig_id = mol.GetProp("lig_id") if mol.HasProp("lig_id") else (
                mol.GetProp("_Name") if mol.HasProp("_Name") else f"pose_{i}"
            )
            if allowed is not None and lig_id not in allowed:
                continue
            row = index.get(lig_id, {})
            for key, val in row.items():
                if key == id_col or val in ("", None):
                    continue
                mol.SetProp(key, str(val))
            writer.write(mol)
    finally:
        writer.close()
    return out_sdf
